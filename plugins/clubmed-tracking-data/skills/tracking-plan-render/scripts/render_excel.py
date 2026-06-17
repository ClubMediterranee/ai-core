#!/usr/bin/env python3
"""
Excel renderer for plan.json — clean layout, no legacy clutter.

Columns:
  A  — N°            (5)
  B  — Événement     (28)   event name, bold navy
  C  — Description   (38)   what this event measures
  D  — Déclencheur   (36)   when it fires (trigger)
  E  — Justification (48)   why this event is tracked (rationale)
  F  — Capture       (20)   screenshot (image if available)
  G  — Push dataLayer(68)   JS code, monospace, wide
  H  — Paramètres    (28)   params list
  I  — Statut        (16)   approved / rejected, color-coded
"""

import json
import sys
from pathlib import Path

LABELS = {
    "en": {
        "title": "Tracking Plan", "sheet": "Specs",
        "number": "#", "event": "Event", "description": "Description",
        "trigger": "Trigger", "justification": "Description",
        "screenshot": "Screenshot", "payload": "Data layer push",
        "params": "Parameters", "status": "Status",
        "approved": "approved event(s)", "open_q": "Open Questions",
        "open_q_title": "Open Questions — events proposed but not included",
        "event_col": "Event", "trigger_col": "Trigger",
        "just_col": "Description", "conf_col": "Confidence",
    },
    "fr": {
        "title": "Plan de marquage", "sheet": "Specs",
        "number": "#", "event": "Événement", "description": "Description",
        "trigger": "Déclencheur", "justification": "Description",
        "screenshot": "Capture", "payload": "Push dataLayer",
        "params": "Paramètres", "status": "Statut",
        "approved": "event(s) approuvé(s)", "open_q": "Open Questions",
        "open_q_title": "Open Questions — events proposés mais non inclus",
        "event_col": "Événement", "trigger_col": "Déclencheur",
        "just_col": "Description", "conf_col": "Confiance",
    },
}

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("ERROR: openpyxl not installed — run: pip3 install openpyxl pillow")
    sys.exit(1)

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY      = "1E2643"
NAVY_SOFT = "2D3F6B"
BLUE_PALE = "EBF2FB"
BLUE_MID  = "D0E4F7"
WHITE     = "FFFFFF"
GREY_HEAD = "4A5568"
GREY_ROW  = "F7F8FA"
CODE_BG   = "F0F4F8"
GREEN     = "C6EFCE"
ORANGE    = "FFEB9C"
RED_LIGHT = "FFC7CE"

def _fill(hex_): return PatternFill("solid", fgColor=hex_)
def _font(color=None, bold=False, mono=False, size=10):
    return Font(
        color=color or "1A1A2E",
        bold=bold,
        name="Courier New" if mono else "Calibri",
        size=size
    )
def _border():
    s = Side(style="thin", color="D0D7E2")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", v="top", wrap=True):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

STATUS_FILL = {
    "approved":        _fill(GREEN),
    "rejected":        _fill(RED_LIGHT),
    "pending_approval":_fill(ORANGE),
}

# ── Column definitions ────────────────────────────────────────────────────────
# (key, header label, col letter, width)
COLUMNS = [
    ("number",          "#",               "A",  5),
    ("event",           "Événement",       "B", 28),
    ("description",     "Description",     "C", 38),
    ("trigger",         "Déclencheur",     "D", 36),
    ("justification",   "Description",   "E", 48),
    ("screenshot",      "Capture",         "F", 20),
    ("payload",         "Push dataLayer",  "G", 68),
    ("params",          "Paramètres",      "H", 28),
    ("status",          "Statut",          "I", 16),
]
COL_LETTER = {key: col for key, _, col, _ in COLUMNS}
COL_IDX    = {key: i+1 for i, (key, _, _, _) in enumerate(COLUMNS)}

# ── Helpers ───────────────────────────────────────────────────────────────────

def build_payload(entry: dict) -> str:
    event   = entry.get("event", "")
    payload = entry.get("payload", {})
    ec      = payload.get("event_click")

    if ec is not None:
        null_push = f'clubMedLayer.push({{\n  "event": "{event}",\n  "event_click": null\n}});\n'
        data      = json.dumps({"event": event, "event_click": ec},
                               ensure_ascii=False, indent=2)
        return null_push + f"\nclubMedLayer.push({data});"
    elif "ecommerce" in payload:
        p = {"event": event, "ecommerce": payload["ecommerce"]}
        return "clubMedLayer.push(\n" + json.dumps(p, ensure_ascii=False, indent=2) + "\n);"
    else:
        p = {"event": event, **{k: v for k, v in payload.items() if k != "event"}}
        return "clubMedLayer.push(\n" + json.dumps(p, ensure_ascii=False, indent=2) + "\n);"


def auto_row_height(payload_text: str, base: int = 20) -> int:
    lines = payload_text.count("\n") + 1
    return max(base, lines * 14 + 8)


def write_title(ws, plan, ncols: int):
    ws.merge_cells(f"A1:{chr(64+ncols)}1")
    c = ws["A1"]
    c.value = f"Plan de marquage — {plan['meta']['name'].upper()}"
    c.fill  = _fill(NAVY)
    c.font  = Font(color=WHITE, bold=True, name="Calibri", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    meta = plan["meta"]
    fields = [
        ("Section",    meta.get("site_section", "")),
        ("TMS",        meta.get("tms", "GTM")),
        ("Analytics",  meta.get("analytics", "GA4")),
        ("Data layer", meta.get("data_layer", "clubMedLayer")),
        ("Généré le",  meta.get("generated_at", "")),
        ("Statut",     meta.get("status", "")),
    ]
    for ci, (label, val) in enumerate(fields, 1):
        lc = ws.cell(row=2, column=ci, value=label)
        lc.fill = _fill(GREY_HEAD); lc.font = _font(WHITE, bold=True)
        lc.alignment = _align("center", "center", False)
        vc = ws.cell(row=3, column=ci, value=val)
        vc.fill = _fill(BLUE_PALE); vc.font = _font()
        vc.alignment = _align("center", "center", False)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 6


def write_headers(ws, row: int):
    ws.row_dimensions[row].height = 32
    for key, label, col, _ in COLUMNS:
        c = ws.cell(row=row, column=COL_IDX[key], value=label)
        c.fill = _fill(NAVY_SOFT)
        c.font = _font(WHITE, bold=True)
        c.alignment = _align("center", "center")
        c.border = _border()


def try_embed_image(ws, entry: dict, row: int, output_dir: str):
    shot = entry.get("screenshot")
    if not shot:
        return
    img_path = Path(output_dir) / shot
    if not img_path.exists():
        ws.cell(row=row, column=COL_IDX["screenshot"], value=shot)\
          .font = _font("0563C1", size=8)
        return
    try:
        img = XLImage(str(img_path))
        img.height, img.width = 90, 120
        ws.add_image(img, f"{COL_LETTER['screenshot']}{row}")
    except Exception:
        ws.cell(row=row, column=COL_IDX["screenshot"], value=shot)\
          .font = _font("0563C1", size=8)


def write_entries(ws, entries: list, output_dir: str, start_row: int = 6):
    row = start_row
    for i, entry in enumerate(entries, 1):
        payload_text = build_payload(entry)
        row_h = auto_row_height(payload_text)
        ws.row_dimensions[row].height = row_h

        # Alternating row background
        row_fill = _fill(BLUE_PALE) if i % 2 == 0 else _fill(GREY_ROW)

        values = {
            "number":        i,
            "event":         entry.get("event", ""),
            "description":   entry.get("description", ""),
            "trigger":       entry.get("trigger", ""),
            "justification": entry.get("rationale", ""),
            "screenshot":    None,
            "payload":       payload_text,
            "params":        "\n".join(
                f"{p['name']} ({p.get('type','string')}): {p.get('description','')}"
                for p in (entry.get("params") or [])
                if isinstance(p, dict)
            ),
            "status":        entry.get("_status", ""),
        }

        for key, _, col, _ in COLUMNS:
            if key == "screenshot":
                continue
            c = ws.cell(row=row, column=COL_IDX[key], value=values[key])
            c.border = _border()

            if key == "payload":
                c.font      = _font(mono=True, size=8)
                c.fill      = _fill(CODE_BG)
                c.alignment = _align("left", "top")
            elif key == "event":
                c.font      = _font(NAVY, bold=True)
                c.fill      = row_fill
                c.alignment = _align("left", "center", False)
            elif key == "number":
                c.font      = _font(bold=True)
                c.fill      = row_fill
                c.alignment = _align("center", "center", False)
            elif key == "status":
                c.fill      = STATUS_FILL.get(entry.get("_status", ""), row_fill)
                c.font      = _font(bold=True, size=9)
                c.alignment = _align("center", "center", False)
            elif key == "justification":
                c.font      = _font("4A5568", size=9)
                c.fill      = row_fill
                c.alignment = _align("left", "top")
            else:
                c.font      = _font()
                c.fill      = row_fill
                c.alignment = _align("left", "top")

        try_embed_image(ws, entry, row, output_dir)
        row += 1

    return row


def render(plan_file: str, output_dir: str, lang: str = "en"):
    plan     = json.load(open(plan_file, encoding="utf-8"))
    name     = plan["meta"]["name"]
    L        = LABELS.get(lang, LABELS["en"])
    approved = [e for e in plan["entries"] if e.get("_status") == "approved"]
    rejected = [e for e in plan["entries"] if e.get("_status") == "rejected"]
    ncols    = len(COLUMNS)

    wb = openpyxl.Workbook()

    # ── Main sheet ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Spécifications techniques"
    ws.freeze_panes = "A6"

    for key, _, col, width in COLUMNS:
        ws.column_dimensions[col].width = width

    write_title(ws, plan, ncols)
    write_headers(ws, row=5)
    last_row = write_entries(ws, approved, output_dir, start_row=6)

    # Summary row
    ws.row_dimensions[last_row].height = 20
    sc = ws.cell(row=last_row, column=2,
                 value=f"Total : {len(approved)} {L['approved']}")
    sc.font = _font(NAVY, bold=True)
    sc.fill = _fill(BLUE_MID)

    # ── Open Questions sheet ───────────────────────────────────────────────────
    if rejected:
        ws2 = wb.create_sheet(L["open_q"])
        ws2.merge_cells("A1:E1")
        t = ws2["A1"]
        t.value = L["open_q_title"]
        t.fill  = _fill(NAVY); t.font = _font(WHITE, bold=True)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        oq_headers = [L["event_col"], L["trigger_col"], L["just_col"], L["conf_col"]]
        oq_widths  = [28, 40, 55, 12]
        for ci, (h, w) in enumerate(zip(oq_headers, oq_widths), 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.fill = _fill(GREY_HEAD); c.font = _font(WHITE, bold=True)
            c.alignment = _align("center", "center")
            ws2.column_dimensions[chr(64+ci)].width = w

        for ri, e in enumerate(rejected, 3):
            ws2.row_dimensions[ri].height = 40
            row_fill = _fill(BLUE_PALE) if ri % 2 == 0 else _fill(GREY_ROW)
            conf = int(float(e.get("confidence", 0)) * 100)
            vals = [e.get("event",""), e.get("trigger",""), e.get("rationale",""), f"{conf}%"]
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.font  = _font(bold=(ci==1))
                c.fill  = row_fill
                c.alignment = _align()
                c.border = _border()

    out_path = Path(output_dir) / f"{name}.xlsx"
    wb.save(str(out_path))
    print(f"✓ Excel written: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: render_excel.py <plan.json> <output_dir> [lang=en|fr]")
        sys.exit(1)
    lang = sys.argv[3] if len(sys.argv) > 3 else "en"
    render(sys.argv[1], sys.argv[2], lang)
